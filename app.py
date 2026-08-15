"""
Lead Finder Web App — Velnex AI
=================================
Streamlit app: pick country + city + business type, scrape via Apify's
Google Maps Scraper, filter/prioritize, and download as Excel.

RUN LOCALLY
-----------
pip install streamlit apify-client openpyxl pandas
streamlit run app.py

DEPLOY
------
Push to GitHub, deploy on Streamlit Community Cloud (same as your
velnexai repo). Add your Apify token as a Streamlit "Secret" instead
of pasting it in the UI every time — see bottom of this file.
"""

import re
import io
import time
import datetime
import requests
import streamlit as st
import pandas as pd
from apify_client import ApifyClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Velnex Lead Finder", page_icon="🎯", layout="wide")

CUSTOM_CSS = """
<style>
:root {
    --navy-dark: #0a2540;
    --blue-deep: #0d3d75;
    --blue-mid: #1a63b8;
    --blue-bright: #2f8fe0;
    --blue-light: #bfe0fb;
}

/* App background gradient */
.stApp {
    background: linear-gradient(160deg, var(--navy-dark) 0%, var(--blue-mid) 55%, var(--blue-light) 100%);
}

/* Decorative network-node motif, bottom right, like the reference image */
.stApp::after {
    content: "";
    position: fixed;
    bottom: 24px;
    right: 24px;
    width: 130px;
    height: 130px;
    pointer-events: none;
    z-index: 0;
    background-image:
        radial-gradient(circle, rgba(255,255,255,0.85) 2px, transparent 2.5px),
        radial-gradient(circle, rgba(255,255,255,0.85) 2px, transparent 2.5px),
        radial-gradient(circle, rgba(255,255,255,0.85) 2px, transparent 2.5px),
        radial-gradient(circle, rgba(255,255,255,0.85) 2px, transparent 2.5px),
        radial-gradient(circle, rgba(255,255,255,0.85) 2px, transparent 2.5px);
    background-position: 10px 10px, 60px 5px, 100px 40px, 40px 70px, 90px 100px;
    background-repeat: no-repeat;
}

/* Title + captions */
h1, h2, h3 {
    color: #ffffff !important;
}
.stCaption, [data-testid="stCaptionContainer"] {
    color: var(--blue-light) !important;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(200deg, var(--navy-dark) 0%, var(--blue-deep) 100%);
}
[data-testid="stSidebar"] * {
    color: #eaf4ff !important;
}

/* Cards / containers for inputs */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-testid="stTextArea"] textarea,
.stSelectbox > div > div {
    background-color: rgba(255,255,255,0.96) !important;
    border-radius: 8px !important;
    border: 1px solid var(--blue-mid) !important;
    color: #0a2540 !important;
    font-weight: 500;
}
[data-testid="stTextInput"] input::placeholder,
[data-testid="stTextArea"] textarea::placeholder {
    color: #5c7fa6 !important;
    opacity: 1 !important;
}

/* Alert / info / warning / success boxes: solid readable background */
[data-testid="stAlert"] {
    background-color: rgba(255,255,255,0.95) !important;
    border-radius: 8px;
}
[data-testid="stAlert"] p, [data-testid="stAlert"] div {
    color: #0a2540 !important;
}

/* Primary button */
.stButton > button[kind="primary"], .stDownloadButton > button[kind="primary"] {
    background: linear-gradient(90deg, var(--blue-deep) 0%, var(--blue-bright) 100%);
    color: white;
    border: none;
    border-radius: 8px;
    font-weight: 600;
}
.stButton > button[kind="primary"]:hover, .stDownloadButton > button[kind="primary"]:hover {
    background: linear-gradient(90deg, var(--blue-mid) 0%, var(--blue-bright) 100%);
    border: none;
}

/* Dataframe container */
[data-testid="stDataFrame"] {
    background: rgba(255,255,255,0.95);
    border-radius: 10px;
    padding: 4px;
}

/* Search button: red, distinct from download button */
.st-key-search_btn_wrap .stButton > button[kind="primary"] {
    background: linear-gradient(90deg, #b3261e 0%, #e03e35 100%) !important;
}
.st-key-search_btn_wrap .stButton > button[kind="primary"]:hover {
    background: linear-gradient(90deg, #931d17 0%, #c9342b 100%) !important;
}

/* Widget labels (Country, City, Business type, etc.) in main area */
[data-testid="stWidgetLabel"] p, [data-testid="stWidgetLabel"] label {
    color: #ffffff !important;
    font-weight: 600;
}

/* Send Emails button: green, distinct from search (red) and download (blue) */
.st-key-send_btn_wrap .stButton > button[kind="primary"] {
    background: linear-gradient(90deg, #1e7d3e 0%, #2fa350 100%) !important;
}
.st-key-send_btn_wrap .stButton > button[kind="primary"]:hover {
    background: linear-gradient(90deg, #17632f 0%, #268a43 100%) !important;
}

/* Tabs styling for visibility on dark background */
.stTabs [data-baseweb="tab"] {
    color: #eaf4ff !important;
    font-weight: 600;
}
.stTabs [aria-selected="true"] {
    color: #ffffff !important;
    border-bottom-color: #2f8fe0 !important;
}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

CHAIN_KEYWORDS_DEFAULT = [
    "aspen dental", "heartland dental", "smile direct", "gentle dental",
    "pacific dental", "western dental", "great expressions", "coast dental",
    "midwest dental", "comfort dental", "supercuts", "great clips",
    "massage envy", "fantastic sams",
]

COUNTRIES = ["United States", "United Kingdom", "Canada", "Pakistan", "Australia", "Other"]

BUSINESS_PRESETS = [
    "Dental clinic", "HVAC contractor", "Law firm", "Hair salon", "Spa",
    "Plumber", "Roofing company", "Physical therapy clinic", "Custom...",
]

COLUMNS = [
    "City", "Business Name", "Phone", "Email", "Website", "Address",
    "Rating", "Review Count", "Google Maps URL", "Category", "Status", "Notes",
]

# ─────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────

def clean_phone(phone: str) -> str:
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[0:3]}) {digits[3:6]}-{digits[6:]}"
    return phone


def extract_email(item: dict) -> str:
    # Different actor versions expose this differently; check common shapes
    emails = item.get("emails")
    if isinstance(emails, list) and emails:
        return emails[0]
    if isinstance(emails, str) and emails:
        return emails
    single = item.get("email") or item.get("emailAddress")
    if single:
        return single
    return ""


def is_chain(name: str, chain_keywords: list) -> bool:
    name_l = name.lower()
    return any(chain in name_l for chain in chain_keywords)


def scrape_city(client: ApifyClient, search_term: str, location: str, max_results: int, find_emails: bool, progress_cb=None):
    run_input = {
        "searchStringsArray": [search_term],
        "locationQuery": location,
        "maxCrawledPlacesPerSearch": max_results,
        "language": "en",
        "skipClosedPlaces": True,
        "scrapeContacts": find_emails,  # pulls emails/names/job titles from each business's website — costs extra per place
    }
    run = client.actor("compass/crawler-google-places").call(run_input=run_input)

    # apify-client versions differ: some return a dict, some return a Run object
    if isinstance(run, dict):
        dataset_id = run["defaultDatasetId"]
    else:
        dataset_id = getattr(run, "default_dataset_id", None) or getattr(run, "defaultDatasetId", None)

    items = list(client.dataset(dataset_id).iterate_items())
    return items


def process_items(items: list, city: str, min_reviews: int, min_rating: float, chain_keywords: list) -> list:
    leads = []
    for item in items:
        name = item.get("title", "").strip()
        if not name or is_chain(name, chain_keywords):
            continue
        reviews = item.get("reviewsCount") or 0
        rating = item.get("totalScore") or 0
        if reviews < min_reviews or rating < min_rating:
            continue
        leads.append({
            "City": city,
            "Business Name": name,
            "Phone": clean_phone(item.get("phone", "")),
            "Email": extract_email(item),
            "Website": item.get("website", ""),
            "Address": item.get("address", ""),
            "Rating": round(rating, 1),
            "Review Count": reviews,
            "Google Maps URL": item.get("url", ""),
            "Category": item.get("categoryName", ""),
            "Status": "Not Contacted",
            "Notes": "",
        })
    return leads


BREVO_API_URL = "https://api.brevo.com/v3/smtp/email"

DEFAULT_SUBJECT_TEMPLATE = "Quick question about {business_name}'s phone/booking setup"
DEFAULT_BODY_TEMPLATE = """Hi {business_name} team,

I'm {from_name}, founder of {business_company}. I build AI voice and chat agents that handle appointment booking, missed-call text-backs, and after-hours inquiries for businesses like yours — so you never lose a customer to a call that went to voicemail.

Would you be open to a quick 10-minute call to see if it's a fit for {business_name}?

Best,
{from_name}
{business_company}
{phone_number}

---
{business_company} | {mailing_address}
Don't want future emails like this? Reply "unsubscribe" and we'll remove you immediately.
"""


def send_via_brevo(api_key: str, from_name: str, from_email: str, to_email: str, to_name: str, subject: str, body: str) -> tuple:
    payload = {
        "sender": {"name": from_name, "email": from_email},
        "to": [{"email": to_email, "name": to_name}],
        "replyTo": {"email": from_email, "name": from_name},
        "subject": subject,
        "textContent": body,
    }
    headers = {"accept": "application/json", "api-key": api_key, "content-type": "application/json"}
    try:
        response = requests.post(BREVO_API_URL, json=payload, headers=headers, timeout=15)
        if response.status_code in (200, 201):
            return True, ""
        return False, f"HTTP {response.status_code}: {response.text[:200]}"
    except Exception as e:
        return False, str(e)


def send_via_gmail_smtp(smtp_email: str, app_password: str, from_name: str, to_email: str, subject: str, body: str) -> tuple:
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    msg = MIMEMultipart()
    msg["From"] = f"{from_name} <{smtp_email}>"
    msg["To"] = to_email
    msg["Subject"] = subject
    msg["Reply-To"] = smtp_email
    msg.attach(MIMEText(body, "plain"))

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(smtp_email, app_password)
            server.sendmail(smtp_email, to_email, msg.as_string())
        return True, ""
    except smtplib.SMTPAuthenticationError:
        return False, "Auth failed — check your App Password (not your normal Google password)."
    except Exception as e:
        return False, str(e)


def build_excel(leads: list) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=10)
    status_options = '"Not Contacted,Emailed,Called,Demo Booked,Demo Done,Won,Lost,Not Interested"'

    ws = wb.create_sheet("All Leads")
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    leads_sorted = sorted(leads, key=lambda r: (r["Review Count"], r["Rating"]), reverse=True)
    for r_idx, row in enumerate(leads_sorted, start=2):
        for c_idx, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(col, ""))
            cell.font = body_font

    last_row = len(leads_sorted) + 1
    widths = [16, 30, 16, 26, 28, 34, 9, 12, 34, 20, 16, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if last_row > 1:
        dv = DataValidation(type="list", formula1=status_options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"K2:K{last_row}")

        ws.conditional_formatting.add(
            f"G2:G{last_row}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                            mid_type="percentile", mid_value=50, mid_color="FFEB84",
                            end_type="max", end_color="63BE7B"),
        )
        ws.conditional_formatting.add(
            f"K2:K{last_row}",
            CellIsRule(operator="equal", formula=['"Won"'],
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
        )
        ws.conditional_formatting.add(
            f"K2:K{last_row}",
            CellIsRule(operator="equal", formula=['"Lost"'],
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────

st.title("🎯 Velnex Lead Finder & Outreach")
st.caption("Find leads, then send them personalized emails — all in one place.")

with st.sidebar:
    st.header("Apify Connection")
    try:
        default_apify_token = st.secrets["APIFY_API_TOKEN"]
    except Exception:
        default_apify_token = ""
    api_token = st.text_input("Apify API Token", value=default_apify_token, type="password",
                               help="console.apify.com → Settings → Integrations")

    st.header("Filters")
    find_emails = st.checkbox(
        "Find emails (costs extra Apify credits)",
        value=False,
        help="Visits each business's website to find an email. Roughly 2-3x the cost per lead compared to phone/address only.",
    )
    min_reviews = st.slider("Minimum reviews", 0, 200, 25)
    min_rating = st.slider("Minimum rating", 0.0, 5.0, 3.8, 0.1)
    exclude_chains = st.checkbox("Exclude known chains", value=True)
    max_results = st.slider("Max results per city", 10, 120, 60)

    st.header("Email Sending Method")
    send_method = st.radio("Send via", ["Brevo (API)", "Gmail/Workspace SMTP"], index=0)

    if send_method == "Brevo (API)":
        try:
            default_brevo_key = st.secrets["BREVO_API_KEY"]
        except Exception:
            default_brevo_key = ""
        brevo_api_key = st.text_input("Brevo API Key", value=default_brevo_key, type="password",
                                       help="Brevo → SMTP & API → API Keys")
        try:
            default_from_email = st.secrets["FROM_EMAIL"]
        except Exception:
            default_from_email = "info@velnexai.com"
        from_email = st.text_input("Sender email", value=default_from_email)
        smtp_app_password = ""
    else:
        try:
            default_from_email = st.secrets["FROM_EMAIL"]
        except Exception:
            default_from_email = "info@velnexai.com"
        from_email = st.text_input("Workspace email", value=default_from_email)
        try:
            default_smtp_pw = st.secrets["GMAIL_APP_PASSWORD"]
        except Exception:
            default_smtp_pw = ""
        smtp_app_password = st.text_input("App Password", value=default_smtp_pw, type="password",
                                           help="myaccount.google.com/apppasswords (16-character code)")
        brevo_api_key = ""

    from_name = st.text_input("Sender name", value="Aashir")
    business_company = st.text_input("Company name (shown in email)", value="Velnex AI")
    phone_number = st.text_input("Phone number (shown in email)", value="")
    mailing_address = st.text_area("Mailing address (legally required in emails)", value="", height=70)

if "leads" not in st.session_state:
    st.session_state.leads = []

tab_find, tab_send = st.tabs(["🔍 Find Leads", "📧 Send Emails"])

# ─────────────────────────────────────────────────────────────────────────
# TAB 1: FIND LEADS
# ─────────────────────────────────────────────────────────────────────────

with tab_find:
    col1, col2, col3 = st.columns(3)
    with col1:
        country = st.selectbox("Country", COUNTRIES)
    with col2:
        cities_input = st.text_input("City / cities (comma-separated)", placeholder="Austin TX, Dallas TX")
    with col3:
        business_preset = st.selectbox("Business type", BUSINESS_PRESETS)
        if business_preset == "Custom...":
            business_type = st.text_input("Enter business type", placeholder="e.g. orthodontist")
        else:
            business_type = business_preset

    with st.container(key="search_btn_wrap"):
        run_button = st.button("🔍 Find Leads", type="primary", use_container_width=True)

    if find_emails:
        st.caption("📧 Email lookup is ON — this run will cost more Apify credits per lead.")
    else:
        st.caption("📧 Email lookup is OFF — faster and cheaper, phone/website/address only.")

    if run_button:
        if not api_token:
            st.error("Enter your Apify API token in the sidebar first.")
        elif not cities_input.strip():
            st.error("Enter at least one city.")
        elif not business_type.strip():
            st.error("Enter or select a business type.")
        else:
            cities = [c.strip() for c in cities_input.split(",") if c.strip()]
            locations = [f"{c}, {country}" if country != "Other" else c for c in cities]

            client = ApifyClient(api_token)
            all_leads = []
            progress = st.progress(0, text="Starting...")
            status = st.empty()

            for i, (city, location) in enumerate(zip(cities, locations)):
                status.text(f"Scraping {city}...")
                try:
                    items = scrape_city(client, business_type, location, max_results, find_emails)
                    chain_list = CHAIN_KEYWORDS_DEFAULT if exclude_chains else []
                    leads = process_items(items, city, min_reviews, min_rating, chain_list)
                    all_leads.extend(leads)
                    status.text(f"{city}: {len(leads)} qualified leads found")
                except Exception as e:
                    st.warning(f"Error scraping {city}: {e}")
                progress.progress((i + 1) / len(cities))

            progress.empty()
            status.empty()
            st.session_state.leads = all_leads
            st.success(f"Done — {len(all_leads)} qualified leads across {len(cities)} location(s). Head to the 'Send Emails' tab to reach out.")

    if st.session_state.leads:
        df = pd.DataFrame(st.session_state.leads).sort_values(
            by=["Review Count", "Rating"], ascending=False
        )
        st.dataframe(df, use_container_width=True, hide_index=True)

        excel_bytes = build_excel(st.session_state.leads)
        filename = f"leads_{business_type.replace(' ', '_') if 'business_type' in dir() else 'export'}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            "⬇️ Download Excel CRM",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
    else:
        st.info("No leads yet — run a search above to get started.")

# ─────────────────────────────────────────────────────────────────────────
# TAB 2: SEND EMAILS
# ─────────────────────────────────────────────────────────────────────────

with tab_send:
    st.caption("This is fully separate from 'Find Leads' — add businesses manually here, write your own script, and send.")

    if "manual_leads_df" not in st.session_state:
        st.session_state.manual_leads_df = pd.DataFrame(
            [{"Business Name": "", "Email": "", "Status": "Pending"}] * 5
        )
    if "manual_email_log" not in st.session_state:
        st.session_state.manual_email_log = {}  # email -> "sent" | "failed"

    st.subheader("1. Add your businesses")
    st.caption("Type or paste directly into the table. Use the + at the bottom to add rows, or right-click to delete.")

    edited_df = st.data_editor(
        st.session_state.manual_leads_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Business Name": st.column_config.TextColumn("Business Name", required=True),
            "Email": st.column_config.TextColumn("Email", required=True),
            "Status": st.column_config.TextColumn("Status", disabled=True),
        },
        key="manual_leads_editor",
    )
    st.session_state.manual_leads_df = edited_df

    st.subheader("2. Write your script")
    st.caption('Use {business_name} anywhere you want it auto-filled — e.g. "Dear {business_name} team,"')

    subject_template = st.text_input("Subject line", value=DEFAULT_SUBJECT_TEMPLATE)
    body_template = st.text_area("Email body", value=DEFAULT_BODY_TEMPLATE, height=280)

    valid_rows = edited_df[(edited_df["Business Name"].str.strip() != "") & (edited_df["Email"].str.strip() != "")]
    valid_leads = valid_rows.to_dict("records")

    daily_limit = st.slider("Max emails to send this run (safety cap)", 1, 300, 90)

    if valid_leads:
        sample = valid_leads[0]
        try:
            preview_subject = subject_template.format(business_name=sample["Business Name"])
            preview_body = body_template.format(
                business_name=sample["Business Name"],
                from_name=from_name,
                business_company=business_company,
                phone_number=phone_number,
                mailing_address=mailing_address,
            )
            with st.expander(f"Preview first email → {sample['Business Name']} ({sample['Email']})"):
                st.text(f"Subject: {preview_subject}\n\n{preview_body}")
        except KeyError as e:
            st.error(f"Your template uses a placeholder I don't recognize: {e}. Supported: {{business_name}}, {{from_name}}, {{business_company}}, {{phone_number}}, {{mailing_address}}")

    st.subheader("3. Send")
    not_sent = [l for l in valid_leads if st.session_state.manual_email_log.get(l["Email"]) != "sent"]
    already_sent = len(valid_leads) - len(not_sent)
    st.caption(f"{len(valid_leads)} valid rows · {already_sent} already sent this session · {len(not_sent)} pending")

    with st.container(key="send_btn_wrap"):
        send_button = st.button("📧 Send Emails", type="primary", use_container_width=True)

    if send_button:
        if send_method == "Brevo (API)" and not brevo_api_key:
            st.error("Enter your Brevo API key in the sidebar first.")
        elif send_method == "Gmail/Workspace SMTP" and not smtp_app_password:
            st.error("Enter your Gmail/Workspace App Password in the sidebar first.")
        elif not from_email or not mailing_address.strip():
            st.error("Sender email and mailing address are required (legal requirement for outreach emails).")
        elif not valid_leads:
            st.error("Add at least one business name + email in the table above.")
        elif not not_sent:
            st.info("Everyone in the table has already been emailed this session.")
        else:
            batch = not_sent[:daily_limit]
            progress = st.progress(0, text="Starting...")
            status = st.empty()
            sent_count, fail_count = 0, 0

            for i, lead in enumerate(batch):
                try:
                    subject = subject_template.format(business_name=lead["Business Name"])
                    body = body_template.format(
                        business_name=lead["Business Name"],
                        from_name=from_name,
                        business_company=business_company,
                        phone_number=phone_number,
                        mailing_address=mailing_address,
                    )
                except KeyError as e:
                    st.error(f"Template error on {lead['Business Name']}: unrecognized placeholder {e}")
                    break

                if send_method == "Brevo (API)":
                    success, error = send_via_brevo(
                        brevo_api_key, from_name, from_email,
                        lead["Email"], lead["Business Name"], subject, body,
                    )
                else:
                    success, error = send_via_gmail_smtp(
                        from_email, smtp_app_password, from_name,
                        lead["Email"], subject, body,
                    )

                if success:
                    sent_count += 1
                    st.session_state.manual_email_log[lead["Email"]] = "sent"
                    status.text(f"Sent -> {lead['Business Name']} ({lead['Email']})")
                else:
                    fail_count += 1
                    st.session_state.manual_email_log[lead["Email"]] = "failed"
                    status.text(f"FAILED -> {lead['Business Name']}: {error}")

                progress.progress((i + 1) / len(batch))
                time.sleep(1)

            progress.empty()
            status.empty()
            st.success(f"Done — Sent: {sent_count} | Failed: {fail_count}")

            for i, row in st.session_state.manual_leads_df.iterrows():
                email = row["Email"]
                if email in st.session_state.manual_email_log:
                    st.session_state.manual_leads_df.at[i, "Status"] = st.session_state.manual_email_log[email]
            st.rerun()
